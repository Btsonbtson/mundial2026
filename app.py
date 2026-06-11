import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
import os

# Ρύθμιση Σελίδας
st.set_page_config(page_title="Mundial 2026 Predictor App", page_icon="⚽", layout="wide")

# CSS Branding (FIFA Theme)
st.markdown("""
    <style>
    .main { background-color: #0f172a; color: #f8fafc; }
    h1, h2, h3 { color: #f59e0b !important; font-family: 'Arial Black', sans-serif; }
    .stButton>button { background-color: #1e3a8a; color: white; border-radius: 8px; border: 1px solid #f59e0b; width: 100%; font-weight: bold; }
    .stButton>button:hover { background-color: #3b82f6; color: white; }
    .chat-box { background-color: #1e293b; padding: 15px; border-radius: 10px; border-left: 5px solid #f59e0b; max-height: 250px; overflow-y: auto; }
    .odds-badge { background-color: #047857; color: white; padding: 3px 8px; border-radius: 5px; font-size: 12px; font-weight: bold; }
    .group-container { background-color: #1e293b; padding: 20px; border-radius: 10px; border: 1px solid #334155; }
    .login-container { max-width: 450px; margin: 100px auto; padding: 30px; background-color: #1e293b; border-radius: 10px; border: 2px solid #f59e0b; }
    </style>
""", unsafe_allow_html=True)

# Εντοπισμός Αρχείου Excel
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(CURRENT_DIR, "Mundial 2026.xlsx")

# Λίστα Κωδικών Πρόσβασης
PASSWORDS = {
    "1453": "BOIKOS",
    "1821": "MAVROMICHALIS",
    "1940": "CHOUSIADAS"
}

# Αρχικοποίηση Session States
if 'logged_in_user' not in st.session_state:
    st.session_state.logged_in_user = None
if 'chat_history' not in st.session_state:
    st.session_state.chat_history = [{"user": "System", "msg": "Το chat ενεργοποιήθηκε με ασφάλεια!"}]

# ΠΡΑΓΜΑΤΙΚΕΣ ΗΜΕΡΟΜΗΝΙΕΣ ΕΝΑΡΞΗΣ MUNDIAL 2026
MATCH_TIMES = {
    5: datetime(2026, 6, 11, 22, 0, 0),  # MEXICO - SOUTH AFRICA
    6: datetime(2026, 6, 12, 18, 0, 0),  # SOUTH KOREA - CZECHIA
    7: datetime(2026, 6, 12, 21, 0, 0),  # CANADA - BOSNIA & HERZ.
    8: datetime(2026, 6, 13, 0, 0, 0),   # SPAIN - ECUADOR
}

STOIXIMAN_ODDS = {
    5: {"1": "1.85", "X": "3.40", "2": "4.50"},
    6: {"1": "2.10", "X": "3.20", "2": "3.60"},
    7: {"1": "1.55", "X": "3.90", "2": "6.00"},
    8: {"1": "1.70", "X": "3.60", "2": "5.20"},
}

# Φόρτωση δεδομένων
wb_read = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
sheet_read = wb_read.active


# --- ΟΘΟΝΗ ΕΙΣΟΔΟΥ (LOGIN PANEL) ---
if st.session_state.logged_in_user is None:
    st.markdown("<div class='login-container'>", unsafe_allow_html=True)
    st.write("## 🔒 Είσοδος στο Mundial Predictor")
    input_pass = st.text_input("Εισάγετε τον Προσωπικό σας Κωδικό:", type="password")
    
    if st.button("Είσοδος 🚀"):
        if input_pass in PASSWORDS:
            st.session_state.logged_in_user = PASSWORDS[input_pass]
            st.success(f"Καλώς ορίσατε, {st.session_state.logged_in_user}!")
            st.rerun()
        else:
            st.error("❌ Λάθος Κωδικός! Προσπαθήστε ξανά.")
    st.markdown("</div>", unsafe_allow_html=True)
    st.stop()  # Σταματάει την εκτέλεση εδώ αν δεν έχει γίνει login


# --- ΚΥΡΙΩΣ ΕΦΑΡΜΟΓΗ (ΜΟΝΟ ΓΙΑ ΣΥΝΔΕΔΕΜΕΝΟΥΣ) ---
st.title("🏆 MUNDIAL 2026 – AUTOMATED PREDICTOR")
st.subheader(f"👤 Συνδεδεμένος Παίκτης: {st.session_state.logged_in_user}")

# Δομή για προσωρινή αποθήκευση αλλαγών (Buffer) πριν το Logout
if 'temp_matches' not in st.session_state:
    st.session_state.temp_matches = {}
if 'temp_groups' not in st.session_state:
    st.session_state.temp_groups = {}

# Κουμπί Έξοδος & Αυτόματη Αποθήκευση
if st.button("🚪 ΕΞΟΔΟΣ & ΑΥΤΟΜΑΤΗ ΑΠΟΘΗΚΕΥΣΗ ΣΤΟ EXCEL"):
    wb_save = openpyxl.load_workbook(EXCEL_FILE)
    sh_save = wb_save.active
    
    # 1. Σώζουμε τους αγώνες από το buffer
    USER_COLS = {"BOIKOS": {"h": "N", "a": "P"}, "MAVROMICHALIS": {"h": "U", "a": "W"}, "CHOUSIADAS": {"h": "AB", "a": "AD"}}
    cols = USER_COLS[st.session_state.logged_in_user]
    for r_idx, scores in st.session_state.temp_matches.items():
        sh_save[f"{cols['h']}{r_idx}"] = scores["h"]
        sh_save[f"{cols['a']}{r_idx}"] = scores["a"]
        
    # 2. Σώζουμε τους ομίλους από το buffer
    PLAYER_OFFSETS = {"BOIKOS": 11, "MAVROMICHALIS": 17, "CHOUSIADAS": 23}
    start_row = PLAYER_OFFSETS[st.session_state.logged_in_user]
    for col_idx, positions in st.session_state.temp_groups.items():
        sh_save.cell(row=start_row, column=col_idx + 1).value = positions[0]
        sh_save.cell(row=start_row+1, column=col_idx + 1).value = positions[1]
        sh_save.cell(row=start_row+2, column=col_idx + 1).value = positions[2]
        sh_save.cell(row=start_row+3, column=col_idx + 1).value = positions[3]
        
    wb_save.save(EXCEL_FILE)
    
    # Καθαρισμός και Reset
    st.session_state.temp_matches = {}
    st.session_state.temp_groups = {}
    st.session_state.logged_in_user = None
    st.success("Όλες οι αλλαγές αποθηκεύτηκαν με ασφάλεια στο Excel!")
    st.rerun()

st.write("---")

# TABS
tab1, tab2 = st.tabs(["⚽ Αγώνες & Σκορ", "📊 Κατάταξη Ομίλων"])

# --- TAB 1: ΑΓΩΝΕΣ ---
with tab1:
    st.header("Προβλέψεις Σκορ Αγώνων")
    USER_COLS = {"BOIKOS": {"h": "N", "a": "P"}, "MAVROMICHALIS": {"h": "U", "a": "W"}, "CHOUSIADAS": {"h": "AB", "a": "AD"}}
    cols = USER_COLS[st.session_state.logged_in_user]
    
    for r in range(5, 9):
        h_team = sheet_read[f"E{r}"].value
        a_team = sheet_read[f"G{r}"].value
        if not h_team: continue
        
        m_time = MATCH_TIMES.get(r, datetime(2026, 6, 15, 12, 0, 0))
        time_to_start = m_time - datetime.now()
        is_locked = time_to_start.total_seconds() <= 60
        
        # Αν δεν υπάρχει ήδη στο buffer, διαβάζουμε την τιμή από το Excel
        if r not in st.session_state.temp_matches:
            st.session_state.temp_matches[r] = {
                "h": sheet_read[f"{cols['h']}{r}"].value or 0,
                "a": sheet_read[f"{cols['a']}{r}"].value or 0
            }
            
        with st.container():
            c1, c2, c3, c4 = st.columns([3, 2, 4, 3])
            with c1:
                st.markdown(f"**{h_team} vs {a_team}**")
                odds = STOIXIMAN_ODDS.get(r, {"1":"-","X":"-","2":"-"})
                st.markdown(f"<span class='odds-badge'>Stoiximan</span> 1:{odds['1']} X:{odds['X']} 2:{odds['2']}", unsafe_allow_html=True)
            with c2:
                st.write(f"📅 Σέντρα: {m_time.strftime('%d/%m %H:%M')}")
                if is_locked: st.error("🔒 Κλειδωμένο")
                else: st.info(f"⏳ Σε: {time_to_start.days}ημ. {time_to_start.seconds // 3600}ώρ.")
            with c3:
                cc1, cc2 = st.columns(2)
                with cc1: 
                    new_h = st.number_input(f"Σκορ {h_team}", 0, 10, int(st.session_state.temp_matches[r]["h"]), key=f"h_{r}", disabled=is_locked)
                with cc2: 
                    new_a = st.number_input(f"Σκορ {a_team}", 0, 10, int(st.session_state.temp_matches[r]["a"]), key=f"a_{r}", disabled=is_locked)
                st.session_state.temp_matches[r] = {"h": new_h, "a": new_a}
            with c4:
                st.write("👀 Αποκάλυψη:")
                if is_locked:
                    st.write(f"BOIKOS: {sheet_read[f'N{r}'].value or 0}-{sheet_read[f'P{r}'].value or 0}")
                    st.write(f"MAVRO: {sheet_read[f'U{r}'].value or 0}-{sheet_read[f'W{r}'].value or 0}")
                    st.write(f"CHOUS: {sheet_read[f'AB{r}'].value or 0}-{sheet_read[f'AD{r}'].value or 0}")
                else:
                    st.warning("🔒 Κρυφό μέχρι το 1'")
        st.markdown("---")

# --- TAB 2: ΟΜΙΛΟΙ ---
with tab2:
    st.header("📊 Προβλέψεις Τελικής Κατάταξης Ομίλων")
    
    GROUPS_MAP = {
        "GROUP A": 37, "GROUP B": 42, "GROUP C": 47, "GROUP D": 52,
        "GROUP E": 57, "GROUP F": 62, "GROUP G": 67, "GROUP H": 72,
        "GROUP I": 77, "GROUP J": 82, "GROUP K": 87, "GROUP L": 92
    }
    
    selected_group = st.selectbox("🗂️ Επιλέξτε Όμιλο για πρόβλεψη:", list(GROUPS_MAP.keys()))
    col_letter_idx = GROUPS_MAP[selected_group]
    
    world_cup_start = datetime(2026, 6, 11, 22, 0, 0)
    time_to_lock = world_cup_start - datetime.now()
    group_locked = time_to_lock.total_seconds() <= 0
    
    if group_locked: st.error("🔒 Οι προβλέψεις ομίλων έχουν κλειδώσει!")
    else: st.info(f"⏳ Χρόνος μέχρι το κλείδωμα: {time_to_lock.days} ημέρες, {time_to_lock.seconds // 3600} ώρες")
    
    PLAYER_OFFSETS = {"BOIKOS": 11, "MAVROMICHALIS": 17, "CHOUSIADAS": 23}
    start_row = PLAYER_OFFSETS[st.session_state.logged_in_user]
    
    group_teams = []
    for r in range(5, 9):
        val = sheet_read.cell(row=r, column=col_letter_idx + 1).value
        if val: group_teams.append(str(val).strip())
        
    if len(group_teams) >= 4:
        if col_letter_idx not in st.session_state.temp_groups:
            st.session_state.temp_groups[col_letter_idx] = [
                sheet_read.cell(row=start_row, column=col_letter_idx + 1).value or group_teams[0],
                sheet_read.cell(row=start_row+1, column=col_letter_idx + 1).value or group_teams[1],
                sheet_read.cell(row=start_row+2, column=col_letter_idx + 1).value or group_teams[2],
                sheet_read.cell(row=start_row+3, column=col_letter_idx + 1).value or group_teams[3]
            ]
            
        g_preds = st.session_state.temp_groups[col_letter_idx]
        
        with st.container():
            st.markdown("<div class='group-container'>", unsafe_allow_html=True)
            p1 = st.selectbox("1η Θέση (Πρόκριση):", group_teams, index=group_teams.index(g_preds[0]) if g_preds[0] in group_teams else 0, disabled=group_locked, key=f"g1_{selected_group}")
            p2 = st.selectbox("2η Θέση (Πρόκριση):", group_teams, index=group_teams.index(g_preds[1]) if g_preds[1] in group_teams else 1, disabled=group_locked, key=f"g2_{selected_group}")
            p3 = st.selectbox("3η Θέση:", group_teams, index=group_teams.index(g_preds[2]) if g_preds[2] in group_teams else 2, disabled=group_locked, key=f"g3_{selected_group}")
            p4 = st.selectbox("4η Θέση:", group_teams, index=group_teams.index(g_preds[3]) if g_preds[3] in group_teams else 3, disabled=group_locked, key=f"g4_{selected_group}")
            
            if len({p1, p2, p3, p4}) < 4:
                st.error("❌ Προσοχή: Έχετε επιλέξει την ίδια ομάδα σε παραπάνω από μία θέσεις!")
            else:
                st.session_state.temp_groups[col_letter_idx] = [p1, p2, p3, p4]
            st.markdown("</div>", unsafe_allow_html=True)

# --- 📊 LIVE LEADERBOARD ---
st.write("---")
st.header("📊 Live Βαθμολογία (Από τις φόρμουλες του Excel)")
score_boikos, score_mavro, score_chous = 0, 0, 0

for r in range(5, 9):
    h_val = sheet_read[f"I{r}"].value
    has_real_score = (h_val is not None) and (str(h_val).strip() != "-") and (str(h_val).strip() != "")
    if has_real_score:
        score_boikos += float(sheet_read[f"S{r}"].value or 0)
        score_mavro += float(sheet_read[f"Z{r}"].value or 0)
        score_chous += float(sheet_read[f"AG{r}"].value or 0)

leaderboard_data = {"Παίκτης": ["BOIKOS", "MAVROMICHALIS", "CHOUSIADAS"], "Συνολικοί Πόντοι": [score_boikos, score_mavro, score_chous]}
df_lb = pd.DataFrame(leaderboard_data).sort_values(by="Συνολικοί Πόντοι", ascending=False)
st.table(df_lb)

# --- CHAT ROOM ---
st.write("---")
st.header("💬 Live Chat")
chat_html = "<div class='chat-box'>"
for chat in st.session_state.chat_history:
    chat_html += f"<p><b>[{chat['user']}]:</b> {chat['msg']}</p
